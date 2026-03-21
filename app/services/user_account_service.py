"""
사용자 계정 관리 서비스 — Supabase Admin API 래퍼

관리자가 직원 계정을 생성/초기화하는 흐름:
1. create_auth_user: Supabase Auth 계정 + users 행 동시 생성
2. reset_user_password: 비밀번호 초기화 + must_change_password 설정
3. bulk_create_users: CSV/XLSX 파싱 결과 일괄 등록
4. parse_xlsx_users: XLSX 파일에서 사용자 행 추출
5. parse_xlsx_all: XLSX 5개 시트(조직/본부/팀/사용자/역량) 전체 파싱
6. bulk_setup_org: 조직→본부→팀→사용자→역량 순서 일괄 등록
"""

import logging
import secrets
import uuid

from app.exceptions import TenopAPIError
from app.utils.supabase_client import get_async_client

logger = logging.getLogger(__name__)


async def create_auth_user(
    email: str,
    password: str | None,
    name: str,
    role: str,
    org_id: str,
    team_id: str | None = None,
    division_id: str | None = None,
) -> dict:
    """Supabase Auth 계정 + users 행 동시 생성.

    Returns:
        {"auth_user_id": str, "temp_password": str, "user": dict}
    """
    temp_password = password or secrets.token_urlsafe(12)
    client = await get_async_client()

    # 1) Supabase Auth 계정 생성
    try:
        auth_res = await client.auth.admin.create_user({
            "email": email,
            "password": temp_password,
            "email_confirm": True,
        })
        auth_user = auth_res.user
        if not auth_user:
            raise TenopAPIError("ADMIN_004", "Auth 계정 생성 실패: 응답 없음", 500)
    except TenopAPIError:
        raise
    except Exception as e:
        logger.error(f"Auth 계정 생성 실패: {email} — {e}")
        raise TenopAPIError(
            "ADMIN_004",
            f"사용자 Auth 계정 생성 실패: {str(e)}",
            500,
            {"email": email},
        )

    # 2) users 테이블 행 생성
    user_data = {
        "id": str(auth_user.id),
        "email": email,
        "name": name,
        "role": role,
        "org_id": org_id,
        "must_change_password": True,
    }
    if team_id:
        user_data["team_id"] = team_id
    if division_id:
        user_data["division_id"] = division_id

    try:
        res = await client.table("users").insert(user_data).execute()
        user_row = res.data[0] if res.data else user_data
    except Exception as e:
        # 롤백: Auth 계정 삭제
        try:
            await client.auth.admin.delete_user(str(auth_user.id))
        except Exception:
            logger.error(f"Auth 롤백 실패: {auth_user.id}")
        raise TenopAPIError(
            "ADMIN_004",
            f"사용자 DB 등록 실패: {str(e)}",
            500,
            {"email": email},
        )

    return {
        "auth_user_id": str(auth_user.id),
        "temp_password": temp_password,
        "user": user_row,
    }


async def reset_user_password(
    user_id: str,
    new_password: str | None = None,
) -> dict:
    """비밀번호 초기화 + must_change_password 설정.

    Returns:
        {"user_id": str, "temp_password": str}
    """
    temp_password = new_password or secrets.token_urlsafe(12)
    client = await get_async_client()

    try:
        await client.auth.admin.update_user_by_id(
            user_id,
            {"password": temp_password},
        )
    except Exception as e:
        logger.error(f"비밀번호 초기화 실패: {user_id} — {e}")
        raise TenopAPIError(
            "ADMIN_005",
            f"비밀번호 초기화 실패: {str(e)}",
            500,
            {"user_id": user_id},
        )

    await client.table("users").update(
        {"must_change_password": True}
    ).eq("id", user_id).execute()

    return {"user_id": user_id, "temp_password": temp_password}


def parse_xlsx_users(file_bytes: bytes) -> list[dict]:
    """XLSX 파일에서 사용자 행 추출.

    첫 번째 시트의 헤더에서 이메일/이름/역할 컬럼을 자동 감지.
    또는 '사용자'/'구성원' 시트가 있으면 우선 사용.
    """
    try:
        import openpyxl
    except ImportError:
        raise TenopAPIError(
            "ADMIN_004", "XLSX 처리에 openpyxl이 필요합니다.", 500,
        )

    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 시트 선택: '사용자' or '구성원' 시트 우선, 없으면 첫 시트
    sheet = None
    for sn in wb.sheetnames:
        if "사용자" in sn or "구성원" in sn:
            sheet = wb[sn]
            break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        wb.close()
        return []

    # 헤더 감지
    header = [str(h).strip().lower().replace(" ", "") if h else "" for h in rows[0]]

    def find_col(*candidates: str) -> int | None:
        for i, h in enumerate(header):
            for c in candidates:
                if c in h:
                    return i
        return None

    c_email = find_col("이메일", "email")
    c_name = find_col("이름", "성명", "name")
    c_role = find_col("역할", "role")
    c_team = find_col("팀", "team")
    c_div = find_col("본부", "division")

    if c_email is None or c_name is None:
        wb.close()
        raise TenopAPIError(
            "ADMIN_004",
            "XLSX 헤더에서 '이메일'과 '이름' 컬럼을 찾을 수 없습니다.",
            422,
        )

    # 헤더 다음 행부터 (2행이 '필수/선택' 설명이면 건너뜀)
    start_row = 1
    if len(rows) > 1:
        second_row = rows[1]
        second_vals = [str(v).strip() if v else "" for v in second_row]
        if any(k in " ".join(second_vals) for k in ("필수", "선택", "required", "optional")):
            start_row = 2

    # 역할 매핑 (엑셀 역할 → 시스템 role)
    role_map = {
        "대표이사": "admin", "본부장": "director", "연구소장": "director",
        "실장": "director", "파트너": "director",
        "팀장": "lead",
        "수석": "lead", "책임": "member", "선임": "member",
        "전임": "member", "팀원": "member",
        "사업관리담당": "member", "인사총무담당": "member",
        "총괄기획": "admin", "개발": "member",
        "admin": "admin", "director": "director", "lead": "lead",
        "member": "member", "executive": "executive",
    }

    result = []
    for row in rows[start_row:]:
        email = str(row[c_email]).strip() if row[c_email] else ""
        name = str(row[c_name]).strip() if row[c_name] else ""
        if not email or not name or email == "None":
            continue

        raw_role = str(row[c_role]).strip() if c_role is not None and row[c_role] else "member"
        sys_role = role_map.get(raw_role, "member")

        user = {"email": email, "name": name, "role": sys_role}
        if c_team is not None and row[c_team]:
            user["team_name"] = str(row[c_team]).strip()
        if c_div is not None and row[c_div]:
            user["division_name"] = str(row[c_div]).strip()
        result.append(user)

    wb.close()
    return result


async def bulk_create_users(
    rows: list[dict],
    org_id: str,
) -> dict:
    """CSV/Excel 파싱 결과를 순회하며 계정 일괄 생성.

    Args:
        rows: [{"email": str, "name": str, "role"?: str, "team_id"?: str, "division_id"?: str}, ...]
        org_id: 조직 ID

    Returns:
        {"total": int, "success_count": int, "failed_count": int, "results": list}
    """
    results = []
    success_count = 0
    failed_count = 0

    for row in rows:
        email = row.get("email", "").strip()
        name = row.get("name", "").strip()
        if not email or not name:
            results.append({
                "email": email or "(빈 값)",
                "success": False,
                "error": "이메일 또는 이름이 비어있습니다.",
            })
            failed_count += 1
            continue

        try:
            res = await create_auth_user(
                email=email,
                password=None,
                name=name,
                role=row.get("role", "member"),
                org_id=org_id,
                team_id=row.get("team_id"),
                division_id=row.get("division_id"),
            )
            results.append({
                "email": email,
                "name": name,
                "success": True,
                "temp_password": res["temp_password"],
            })
            success_count += 1
        except TenopAPIError as e:
            results.append({
                "email": email,
                "success": False,
                "error": e.message,
            })
            failed_count += 1
        except Exception as e:
            results.append({
                "email": email,
                "success": False,
                "error": str(e),
            })
            failed_count += 1

    return {
        "total": len(rows),
        "success_count": success_count,
        "failed_count": failed_count,
        "results": results,
    }


# ── 5개 시트 전체 파싱 (조직/본부/팀/사용자/역량) ──────────────────────

def _find_sheet(wb, *keywords):
    """시트 이름에 keyword가 포함된 시트 반환."""
    for sn in wb.sheetnames:
        for kw in keywords:
            if kw in sn:
                return wb[sn]
    return None


def _parse_sheet_rows(sheet) -> list[list]:
    """시트에서 헤더+데이터 행 추출. 2행째가 '필수/선택'이면 스킵."""
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = rows[0]
    start = 1
    if len(rows) > 1:
        second = [str(v).strip() if v else "" for v in rows[1]]
        if any(k in " ".join(second) for k in ("필수", "선택", "required", "optional")):
            start = 2
    return [header] + rows[start:]


def parse_xlsx_all(file_bytes: bytes) -> dict:
    """XLSX 5개 시트(조직/본부/팀/사용자/역량) 전체 파싱.

    Returns:
        {
            "org_name": str | None,
            "divisions": [{"name": str, "org_name": str}],
            "teams": [{"name": str, "division_name": str, "specialty": str | None, "webhook_url": str | None}],
            "users": [{"email": str, "name": str, "role": str, "team_name": str | None, "division_name": str | None}],
            "capabilities": [{"type": str, "title": str, "detail": str, "keywords": list, "org_name": str}],
        }
    """
    try:
        import openpyxl
    except ImportError:
        raise TenopAPIError("ADMIN_004", "XLSX 처리에 openpyxl이 필요합니다.", 500)

    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    result: dict = {
        "org_name": None,
        "divisions": [],
        "teams": [],
        "users": [],
        "capabilities": [],
    }

    # ── 1_조직 ──
    s = _find_sheet(wb, "조직", "org")
    if s:
        rows = _parse_sheet_rows(s)
        if len(rows) >= 2 and rows[1][0]:
            result["org_name"] = str(rows[1][0]).strip()

    # ── 2_본부 ──
    s = _find_sheet(wb, "본부", "division")
    if s:
        rows = _parse_sheet_rows(s)
        for row in rows[1:]:
            name = str(row[0]).strip() if row[0] else ""
            org_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if name:
                result["divisions"].append({"name": name, "org_name": org_name})

    # ── 3_팀 ──
    s = _find_sheet(wb, "팀", "team")
    if s:
        rows = _parse_sheet_rows(s)
        for row in rows[1:]:
            name = str(row[0]).strip() if row[0] else ""
            div_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            specialty = str(row[2]).strip() if len(row) > 2 and row[2] else None
            webhook = str(row[3]).strip() if len(row) > 3 and row[3] else None
            if name:
                result["teams"].append({
                    "name": name,
                    "division_name": div_name,
                    "specialty": specialty,
                    "webhook_url": webhook,
                })

    # ── 4_사용자 ── (기존 parse_xlsx_users 로직 재사용)
    result["users"] = parse_xlsx_users(file_bytes)

    # ── 5_역량 ──
    s = _find_sheet(wb, "역량", "capability")
    if s:
        rows = _parse_sheet_rows(s)
        for row in rows[1:]:
            cap_type = str(row[0]).strip() if row[0] else ""
            title = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            detail = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            kw_str = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            org_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            if cap_type and title:
                keywords = [k.strip() for k in kw_str.split(",") if k.strip()] if kw_str else []
                result["capabilities"].append({
                    "type": cap_type,
                    "title": title,
                    "detail": detail,
                    "keywords": keywords,
                    "org_name": org_name,
                })

    wb.close()
    return result


async def bulk_setup_org(parsed: dict, org_id: str) -> dict:
    """조직→본부→팀→사용자→역량 순서로 일괄 등록.

    Args:
        parsed: parse_xlsx_all() 결과
        org_id: 조직 UUID

    Returns:
        {
            "divisions": {"created": int, "skipped": int},
            "teams": {"created": int, "skipped": int},
            "users": bulk_create_users 결과,
            "capabilities": {"created": int, "skipped": int},
        }
    """
    client = await get_async_client()
    summary: dict = {}

    # ── 1. 본부 등록 (upsert by name) ──
    div_created, div_skipped = 0, 0
    div_name_to_id: dict[str, str] = {}

    # 기존 본부 캐시
    existing = await client.table("divisions").select("id, name").eq("org_id", org_id).execute()
    for d in (existing.data or []):
        div_name_to_id[d["name"]] = d["id"]

    for div in parsed.get("divisions", []):
        name = div["name"]
        if name in div_name_to_id:
            div_skipped += 1
            continue
        new_id = str(uuid.uuid4())
        try:
            await client.table("divisions").insert({
                "id": new_id, "org_id": org_id, "name": name,
            }).execute()
            div_name_to_id[name] = new_id
            div_created += 1
        except Exception as e:
            logger.warning(f"본부 등록 실패: {name} — {e}")
            div_skipped += 1

    summary["divisions"] = {"created": div_created, "skipped": div_skipped}

    # ── 2. 팀 등록 (upsert by name) ──
    team_created, team_skipped = 0, 0
    team_name_to_id: dict[str, str] = {}

    existing = await client.table("teams").select("id, name").execute()
    for t in (existing.data or []):
        team_name_to_id[t["name"]] = t["id"]

    for team in parsed.get("teams", []):
        name = team["name"]
        if name in team_name_to_id:
            team_skipped += 1
            continue
        div_id = div_name_to_id.get(team.get("division_name", ""))
        if not div_id:
            logger.warning(f"팀 '{name}' — 본부 '{team.get('division_name')}' 미발견, 건너뜀")
            team_skipped += 1
            continue
        new_id = str(uuid.uuid4())
        row = {"id": new_id, "division_id": div_id, "name": name}
        if team.get("specialty"):
            row["specialty"] = team["specialty"]
        if team.get("webhook_url"):
            row["webhook_url"] = team["webhook_url"]
        try:
            await client.table("teams").insert(row).execute()
            team_name_to_id[name] = new_id
            team_created += 1
        except Exception as e:
            logger.warning(f"팀 등록 실패: {name} — {e}")
            team_skipped += 1

    summary["teams"] = {"created": team_created, "skipped": team_skipped}

    # ── 3. 사용자 등록 (team_name/division_name → ID 매핑) ──
    for row in parsed.get("users", []):
        if row.get("division_name") and not row.get("division_id"):
            row["division_id"] = div_name_to_id.get(row["division_name"])
        if row.get("team_name") and not row.get("team_id"):
            row["team_id"] = team_name_to_id.get(row["team_name"])

    summary["users"] = await bulk_create_users(parsed.get("users", []), org_id)

    # ── 4. 역량 등록 ──
    cap_created, cap_skipped = 0, 0
    first_user_id = None
    if summary["users"]["results"]:
        for r in summary["users"]["results"]:
            if r.get("success"):
                first_user_id = r.get("auth_user_id")
                break

    for cap in parsed.get("capabilities", []):
        try:
            row = {
                "org_id": org_id,
                "type": cap["type"],
                "title": cap["title"],
                "detail": cap.get("detail", ""),
                "keywords": cap.get("keywords", []),
            }
            if first_user_id:
                row["created_by"] = first_user_id
            await client.table("capabilities").insert(row).execute()
            cap_created += 1
        except Exception as e:
            logger.warning(f"역량 등록 실패: {cap['title']} — {e}")
            cap_skipped += 1

    summary["capabilities"] = {"created": cap_created, "skipped": cap_skipped}

    return summary
