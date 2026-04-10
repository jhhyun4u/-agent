"""
엑셀 팀 구조 → data/team_structure.json 변환 스크립트

사용법:
    uv run python scripts/import_team_structure.py "path/to/tenopa team structure.xlsx"
    uv run python scripts/import_team_structure.py  # 기존 JSON 확인

엑셀 시트 구조 (tenopa team structure.xlsx):
  - 2_본부: 본부명, 소속 조직명
  - 3_팀: 팀명, 소속 본부명, 특화 분야, Teams Webhook URL
  - 4_사용자: 이메일, 이름, 직급, 역할, 소속 팀명, 소속 본부명, 소속 조직명, 관심분야, Azure AD OID
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_PATH = PROJECT_ROOT / "data" / "team_structure.json"


def _cell(row, idx):
    """셀 값 안전 추출."""
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    val = str(row[idx]).strip()
    # 엑셀 오류값 정리
    if val.endswith("'"):
        val = val[:-1]
    return val


def _find_col(header: list[str], *candidates: str) -> int | None:
    """헤더에서 후보 키워드로 컬럼 인덱스 탐색."""
    for i, h in enumerate(header):
        h_lower = h.lower().replace(" ", "")
        for c in candidates:
            if c in h_lower:
                return i
    return None


def import_from_excel(excel_path: str) -> dict:
    """엑셀 파일을 읽어 team_structure 딕셔너리로 변환."""
    try:
        import openpyxl
    except ImportError:
        print("[ERROR] openpyxl 필요: uv add openpyxl --group dev")
        sys.exit(1)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"  시트: {wb.sheetnames}")

    # ── 본부 시트 ──
    division_sheet = None
    for sn in wb.sheetnames:
        if "본부" in sn:
            division_sheet = wb[sn]
            break

    divisions_list = []
    if division_sheet:
        rows = list(division_sheet.iter_rows(values_only=True))
        # 행 2부터 (행 0=헤더, 행 1=필수/선택)
        for row in rows[2:]:
            name = _cell(row, 0)
            if name:
                divisions_list.append(name)
        print(f"  본부: {divisions_list}")

    # ── 팀 시트 ──
    team_sheet = None
    for sn in wb.sheetnames:
        if "팀" in sn:
            team_sheet = wb[sn]
            break

    teams_raw: dict[str, dict] = {}  # team_name → {division, specializations}
    if team_sheet:
        rows = list(team_sheet.iter_rows(values_only=True))
        header = [str(h).strip() if h else "" for h in rows[0]]

        c_name = _find_col(header, "팀명", "team")
        c_div = _find_col(header, "본부", "division")
        c_spec = _find_col(header, "특화", "전문", "specialization", "키워드")

        for row in rows[2:]:  # skip header + 필수/선택 row
            tname = _cell(row, c_name)
            if not tname:
                continue
            div = _cell(row, c_div)
            specs_raw = _cell(row, c_spec)
            specs = [s.strip() for s in specs_raw.split(",") if s.strip()] if specs_raw else []

            teams_raw[tname] = {
                "division": div,
                "specializations": specs,
                "lead": None,
                "members": [],
            }

    # ── 팀명 별칭 매핑 (사용자 시트에서 쓰는 약칭 → 팀 시트 정식 명칭) ──
    team_aliases: dict[str, str] = {}
    # 자동 감지: 같은 본부인데 팀 시트에 없는 이름 → 가장 유사한 팀으로 매핑
    # 수동 힌트 (알려진 패턴)
    alias_hints = {
        "AX1팀": "버티컬AX1팀",
        "T2B1팀": "기술사업화1팀",
        "AX추진팀": "AX혁신팀",
    }
    for short, formal in alias_hints.items():
        if formal in teams_raw:
            team_aliases[short] = formal

    # ── 사용자 시트 ──
    user_sheet = None
    for sn in wb.sheetnames:
        if "사용자" in sn or "구성원" in sn:
            user_sheet = wb[sn]
            break

    users = []
    headquarters_leads = []
    if user_sheet:
        rows = list(user_sheet.iter_rows(values_only=True))
        header = [str(h).strip() if h else "" for h in rows[0]]

        c_email = _find_col(header, "이메일", "email")
        c_name = _find_col(header, "이름", "성명", "name")
        c_grade = _find_col(header, "직급", "grade")
        c_role = _find_col(header, "역할", "role")
        c_team = _find_col(header, "팀명", "team")
        c_div = _find_col(header, "본부명", "division")
        _find_col(header, "조직명", "org")
        c_interest = _find_col(header, "관심", "interest")

        for row in rows[2:]:  # skip header + 필수/선택 row
            name = _cell(row, c_name)
            if not name:
                continue

            user = {
                "email": _cell(row, c_email),
                "name": name,
                "grade": _cell(row, c_grade),
                "role": _cell(row, c_role),
                "team": _cell(row, c_team),
                "division": _cell(row, c_div),
                "interest": _cell(row, c_interest),
            }
            users.append(user)

            team_name = user["team"]
            role = user["role"]

            # 별칭 → 정식명 변환
            canonical_team = team_aliases.get(team_name, team_name)

            # 본부장 판별
            if role == "본부장":
                headquarters_leads.append({
                    "name": name,
                    "division": user["division"],
                    "grade": user["grade"],
                    "role": role,
                })
                continue  # 본부장은 팀 멤버에 추가하지 않음

            # 팀 멤버 등록
            if canonical_team in teams_raw:
                td = teams_raw[canonical_team]
                td["members"].append(name)
                if "팀장" in role or "팀장" in user["grade"]:
                    td["lead"] = name
            elif team_name and team_name not in divisions_list:
                # 엑셀에서 팀 시트에 없는 팀 (경영기획실 등)
                if team_name not in teams_raw:
                    teams_raw[team_name] = {
                        "division": user["division"],
                        "specializations": [],
                        "lead": None,
                        "members": [name],
                    }
                    if "실장" in role or "팀장" in role:
                        teams_raw[team_name]["lead"] = name
                else:
                    teams_raw[team_name]["members"].append(name)
                    if "실장" in role or "팀장" in role:
                        teams_raw[team_name]["lead"] = name

    wb.close()

    # ── 오타 수정 ──
    typo_fixes = {"피지컬먀": "피지컬AI", "재안안전": "재난안전"}
    for tdata in teams_raw.values():
        tdata["specializations"] = [
            typo_fixes.get(s, s) for s in tdata["specializations"]
        ]

    # ── 본부-팀 매핑 구축 ──
    div_teams: dict[str, list[str]] = defaultdict(list)
    for tname, tdata in teams_raw.items():
        div_teams[tdata["division"]].append(tname)

    # 구조화
    result = {
        "organization": "TENOPA",
        "full_name": "테크노베이션파트너스",
        "divisions": [
            {"name": div, "teams": sorted(div_teams.get(div, []))}
            for div in (divisions_list + [d for d in div_teams if d not in divisions_list])
            if div
        ],
        "teams": [
            {
                "name": tname,
                "division": tdata["division"],
                "specializations": tdata["specializations"],
                "lead": tdata["lead"],
                "members": tdata["members"],
            }
            for tname, tdata in teams_raw.items()
        ],
        "headquarters_leads": headquarters_leads,
        "users": users,
    }

    return result


def main():
    parser = argparse.ArgumentParser(description="엑셀 팀 구조 → JSON 변환")
    parser.add_argument("excel_path", nargs="?", default=None, help="엑셀 파일 경로")
    args = parser.parse_args()

    if args.excel_path:
        path = Path(args.excel_path)
        if not path.exists():
            print(f"[ERROR] 파일 없음: {path}")
            sys.exit(1)

        print(f"엑셀 임포트: {path}")
        result = import_from_excel(str(path))
    else:
        # 기존 JSON이 있으면 안내
        if OUTPUT_PATH.exists():
            print(f"기존 파일 있음: {OUTPUT_PATH}")
            with open(OUTPUT_PATH, encoding="utf-8") as f:
                existing = json.load(f)
            n_teams = len(existing.get("teams", []))
            n_users = len(existing.get("users", []))
            n_div = len(existing.get("divisions", []))
            print(f"  {n_div}본부, {n_teams}팀, {n_users}명")
            for t in existing.get("teams", []):
                specs = ", ".join(t["specializations"][:5])
                if len(t["specializations"]) > 5:
                    specs += f" 외 {len(t['specializations']) - 5}개"
                print(f"  - {t['name']} ({t['division']}): {specs or '(없음)'}")
            print("\n엑셀 경로를 지정하면 새로 임포트합니다.")
            return
        else:
            print("[ERROR] 엑셀 경로를 지정하세요.")
            print("  uv run python scripts/import_team_structure.py 'path/to/team.xlsx'")
            sys.exit(1)

    # 저장
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    # 요약
    n_div = len(result["divisions"])
    n_teams = len(result["teams"])
    n_users = len(result["users"])
    total_specs = sum(len(t["specializations"]) for t in result["teams"])

    print(f"\n[OK] {OUTPUT_PATH} 생성 완료")
    print(f"  {n_div}본부, {n_teams}팀, {n_users}명")
    print(f"  전문분야 키워드 총 {total_specs}개")

    for t in result["teams"]:
        specs = ", ".join(t["specializations"][:5])
        if len(t["specializations"]) > 5:
            specs += f" 외 {len(t['specializations']) - 5}개"
        print(f"  - {t['name']} ({t['division']}): {specs or '(없음)'}")


if __name__ == "__main__":
    main()
