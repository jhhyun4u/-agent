"""시드 데이터 입력용 Excel 템플릿 생성 스크립트."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ── 스타일 ──
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 연노랑
OPTIONAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 연초록
EXAMPLE_FONT = Font(name="맑은 고딕", color="808080", italic=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def setup_sheet(ws, columns, examples, note=None):
    """시트 헤더 + 예시 데이터 + 서식 설정."""
    # 헤더
    for col_idx, (name, width, required) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 필수/선택 표시 (2행)
    for col_idx, (name, width, required) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value="필수" if required else "선택")
        cell.fill = REQUIRED_FILL if required else OPTIONAL_FILL
        cell.font = Font(name="맑은 고딕", bold=True, size=9)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # 예시 데이터 (3행~)
    for row_idx, row_data in enumerate(examples, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = EXAMPLE_FONT
            cell.border = THIN_BORDER

    # 안내 메모 (예시 아래 2줄 띄고)
    if note:
        note_row = 3 + len(examples) + 2
        cell = ws.cell(row=note_row, column=1, value=note)
        cell.font = Font(name="맑은 고딕", color="FF0000", size=10)

    ws.freeze_panes = "A3"  # 헤더+필수 행 고정


# ══════════════════════════════════════════════
# 1. 조직 (Organizations)
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "1_조직"
setup_sheet(ws1, [
    ("조직명", 30, True),
], [
    ("TENOPA",),
], note="※ 예시 데이터(회색 이탤릭)는 삭제하고 실제 데이터를 입력하세요.")


# ══════════════════════════════════════════════
# 2. 본부 (Divisions)
# ══════════════════════════════════════════════
ws2 = wb.create_sheet("2_본부")
setup_sheet(ws2, [
    ("본부명", 30, True),
    ("소속 조직명", 25, True),
], [
    ("디지털사업본부", "TENOPA"),
    ("공공사업본부", "TENOPA"),
], note="※ '소속 조직명'은 1_조직 시트에 입력한 이름과 정확히 일치해야 합니다.")


# ══════════════════════════════════════════════
# 3. 팀 (Teams)
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("3_팀")
setup_sheet(ws3, [
    ("팀명", 25, True),
    ("소속 본부명", 25, True),
    ("Teams Webhook URL", 50, False),
], [
    ("AI솔루션팀", "디지털사업본부", "https://outlook.office.com/webhook/..."),
    ("클라우드팀", "디지털사업본부", ""),
    ("행정시스템팀", "공공사업본부", ""),
], note="※ '소속 본부명'은 2_본부 시트에 입력한 이름과 정확히 일치해야 합니다.")


# ══════════════════════════════════════════════
# 4. 사용자 (Users)
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("4_사용자")
cols_user = [
    ("이메일", 30, True),
    ("이름", 15, True),
    ("역할", 12, True),
    ("소속 팀명", 20, False),
    ("소속 본부명", 20, False),
    ("소속 조직명", 20, True),
    ("Azure AD OID", 40, False),
]
setup_sheet(ws4, cols_user, [
    ("admin@tenopa.com", "관리자", "admin", "AI솔루션팀", "디지털사업본부", "TENOPA", ""),
    ("hong@tenopa.com", "홍길동", "lead", "AI솔루션팀", "디지털사업본부", "TENOPA", ""),
    ("kim@tenopa.com", "김철수", "member", "AI솔루션팀", "디지털사업본부", "TENOPA", ""),
    ("park@tenopa.com", "박영희", "member", "클라우드팀", "디지털사업본부", "TENOPA", ""),
    ("lee@tenopa.com", "이사장", "director", "", "디지털사업본부", "TENOPA", ""),
    ("ceo@tenopa.com", "대표이사", "executive", "", "", "TENOPA", ""),
], note="※ 역할: member(팀원), lead(팀장), director(본부장), executive(임원), admin(관리자)")

# 역할 드롭다운
dv_role = DataValidation(type="list", formula1='"member,lead,director,executive,admin"', allow_blank=False)
dv_role.error = "member, lead, director, executive, admin 중 선택하세요."
dv_role.errorTitle = "잘못된 역할"
ws4.add_data_validation(dv_role)
dv_role.add(f"C3:C200")


# ══════════════════════════════════════════════
# 5. 자사 역량 (Capabilities)
# ══════════════════════════════════════════════
ws5 = wb.create_sheet("5_역량")
cols_cap = [
    ("유형", 15, True),
    ("제목", 35, True),
    ("상세 내용", 60, True),
    ("키워드 (쉼표 구분)", 40, False),
    ("소속 조직명", 20, True),
]
setup_sheet(ws5, cols_cap, [
    ("track_record", "클라우드 ERP 구축", "A공공기관 ERP 클라우드 전환 (2024, 10억원 규모)", "ERP, 클라우드, 공공", "TENOPA"),
    ("tech", "AI/ML 파이프라인", "LangGraph 기반 멀티에이전트 아키텍처 구현 역량 보유", "AI, LangGraph, 에이전트", "TENOPA"),
    ("personnel", "PMP 인력 10명", "PMP 자격 보유 PM/PL 10명, 평균 경력 12년", "PM, PMP, 인력", "TENOPA"),
], note="※ 유형: track_record(수행실적), tech(기술역량), personnel(인력)")

dv_type = DataValidation(type="list", formula1='"track_record,tech,personnel"', allow_blank=False)
dv_type.error = "track_record, tech, personnel 중 선택하세요."
ws5.add_data_validation(dv_type)
dv_type.add("A3:A200")


# ── 저장 ──
output_path = "data/seed_template.xlsx"
import os
os.makedirs("data", exist_ok=True)
wb.save(output_path)
print(f"템플릿 생성 완료: {output_path}")
