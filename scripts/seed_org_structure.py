"""
XLSX 기반 조직 구조 초기 세팅 스크립트

tenopa team structure.xlsx 파일을 읽어서:
1. 조직 생성
2. 본부 생성
3. 팀 생성
4. 사용자 생성 (Supabase Auth + users 테이블)
5. 역량 데이터 등록

Usage:
    uv run python scripts/seed_org_structure.py <xlsx_path>
"""

import asyncio
import os
import sys

import openpyxl

# 프로젝트 루트를 sys.path에 추가
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.utils.supabase_client import get_async_client


def parse_xlsx(path: str) -> dict:
    """XLSX 5개 시트 파싱"""
    wb = openpyxl.load_workbook(path)
    result = {"orgs": [], "divisions": [], "teams": [], "users": [], "capabilities": []}

    # 1_조직 (row 3+)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = (row[0] or "").strip()
        if name and not name.startswith("※"):
            result["orgs"].append(name)

    # 2_본부 (row 3+)
    ws = wb[wb.sheetnames[1]]
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = (row[0] or "").strip()
        org = (row[1] or "").strip() if len(row) > 1 else ""
        if name:
            result["divisions"].append({"name": name, "org": org})

    # 3_팀 (row 3+)
    ws = wb[wb.sheetnames[2]]
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = (row[0] or "").strip()
        div = (row[1] or "").strip() if len(row) > 1 else ""
        specialty = (row[2] or "").strip() if len(row) > 2 else ""
        webhook = (row[3] or "").strip() if len(row) > 3 else ""
        if name:
            result["teams"].append({"name": name, "division": div, "specialty": specialty, "webhook": webhook})

    # 4_사용자 (row 3+)
    ws = wb[wb.sheetnames[3]]
    for row in ws.iter_rows(min_row=3, values_only=True):
        email = (row[0] or "").strip()
        uname = (row[1] or "").strip().rstrip("'")  # 박하윤' → 박하윤
        title = (row[2] or "").strip() if len(row) > 2 else ""
        role = (row[3] or "").strip() if len(row) > 3 else ""
        team = (row[4] or "").strip() if len(row) > 4 else ""
        div = (row[5] or "").strip() if len(row) > 5 else ""
        org = (row[6] or "").strip() if len(row) > 6 else ""
        interests = (row[7] or "").strip() if len(row) > 7 else ""
        azure_oid = (row[8] or "").strip() if len(row) > 8 else ""
        if email:
            result["users"].append({
                "email": email, "name": uname, "title": title,
                "role": role, "team": team, "division": div, "org": org,
                "interests": interests, "azure_ad_oid": azure_oid,
            })

    # 5_역량 (row 3+)
    ws = wb[wb.sheetnames[4]]
    for row in ws.iter_rows(min_row=3, values_only=True):
        ctype = (row[0] or "").strip()
        ctitle = (row[1] or "").strip() if len(row) > 1 else ""
        detail = (row[2] or "").strip() if len(row) > 2 else ""
        keywords = (row[3] or "").strip() if len(row) > 3 else ""
        org = (row[4] or "").strip() if len(row) > 4 else ""
        if ctype and ctitle and not ctype.startswith("※"):
            result["capabilities"].append({
                "type": ctype, "title": ctitle, "detail": detail,
                "keywords": keywords, "org": org,
            })

    return result


# XLSX 역할 → DB 역할 매핑
ROLE_MAP = {
    "본부장": "director",
    "팀장": "lead",
    "팀원": "member",
    "총괄기획": "director",
    "개발": "member",
    "실장": "director",
    "사업관리담당": "member",
    "인사총무담당": "member",
}


async def seed(data: dict):
    """DB에 순차 삽입"""
    client = await get_async_client()
    seen_emails = set()

    # ── 1. 조직 ──
    org_map = {}  # name → id
    for org_name in data["orgs"]:
        # 중복 체크
        existing = await client.table("organizations").select("id").eq("name", org_name).limit(1).execute()
        if existing.data and len(existing.data) > 0:
            org_map[org_name] = existing.data[0]["id"]
            print(f"  [조직] '{org_name}' 이미 존재 → {existing.data[0]['id']}")
        else:
            res = await client.table("organizations").insert({"name": org_name}).execute()
            org_map[org_name] = res.data[0]["id"]
            print(f"  [조직] '{org_name}' 생성 → {res.data[0]['id']}")

    # ── 2. 본부 ──
    # XLSX 4_사용자에서 추가 본부 수집 (혁신전략본부, 경영관리본부 등)
    all_div_names = set(d["name"] for d in data["divisions"])
    for u in data["users"]:
        if u["division"] and u["division"] not in all_div_names:
            all_div_names.add(u["division"])
            data["divisions"].append({"name": u["division"], "org": u["org"]})

    div_map = {}  # name → id
    for div in data["divisions"]:
        org_id = org_map.get(div["org"])
        if not org_id:
            print(f"  [본부] '{div['name']}' — 조직 '{div['org']}' 없음, 스킵")
            continue
        existing = await client.table("divisions").select("id").eq("name", div["name"]).eq("org_id", org_id).limit(1).execute()
        if existing.data and len(existing.data) > 0:
            div_map[div["name"]] = existing.data[0]["id"]
            print(f"  [본부] '{div['name']}' 이미 존재")
        else:
            res = await client.table("divisions").insert({"name": div["name"], "org_id": org_id}).execute()
            div_map[div["name"]] = res.data[0]["id"]
            print(f"  [본부] '{div['name']}' 생성")

    # ── 3. 팀 ──
    # XLSX 4_사용자에서 추가 팀 수집
    all_team_names = set(t["name"] for t in data["teams"])
    for u in data["users"]:
        if u["team"] and u["team"] not in all_team_names:
            all_team_names.add(u["team"])
            data["teams"].append({"name": u["team"], "division": u["division"], "specialty": "", "webhook": ""})

    team_map = {}  # name → id
    # division_id → 팀 매핑 (사용자 배정 시 필요)
    team_div_map = {}  # team_name → div_name
    for team in data["teams"]:
        team_div_map[team["name"]] = team["division"]
        existing = await client.table("teams").select("id").eq("name", team["name"]).limit(1).execute()
        if existing.data and len(existing.data) > 0:
            team_map[team["name"]] = existing.data[0]["id"]
            print(f"  [팀] '{team['name']}' 이미 존재")
        else:
            insert_data = {"name": team["name"]}
            # division_id 컬럼이 있으면 추가 (없으면 무시)
            div_id = div_map.get(team["division"])
            try:
                if div_id:
                    insert_data["division_id"] = div_id
                res = await client.table("teams").insert(insert_data).execute()
            except Exception:
                # division_id 컬럼 없으면 없이 재시도
                insert_data = {"name": team["name"]}
                res = await client.table("teams").insert(insert_data).execute()
            team_map[team["name"]] = res.data[0]["id"]
            print(f"  [팀] '{team['name']}' 생성")

    # ── 4. 사용자 ──
    org_id = list(org_map.values())[0] if org_map else None
    for u in data["users"]:
        if u["email"] in seen_emails:
            print(f"  [사용자] '{u['email']}' 중복 스킵 (겸직)")
            continue
        seen_emails.add(u["email"])

        team_id = team_map.get(u["team"])
        div_id = div_map.get(u["division"])
        role = ROLE_MAP.get(u["role"], "member")

        # users 테이블에 이미 있는지 확인
        existing = await client.table("users").select("id").eq("email", u["email"]).limit(1).execute()
        if existing.data and len(existing.data) > 0:
            # 업데이트 (소속 변경 반영)
            updates = {"name": u["name"], "role": role}
            if team_id:
                updates["team_id"] = team_id
            if div_id:
                updates["division_id"] = div_id
            await client.table("users").update(updates).eq("id", existing.data[0]["id"]).execute()
            print(f"  [사용자] '{u['email']}' 이미 존재 → 업데이트")
        else:
            # Supabase Auth 계정 생성 + users 행 삽입
            try:
                from app.services import user_account_service
                result = await user_account_service.create_auth_user(
                    email=u["email"],
                    password=None,  # 자동 임시 비밀번호
                    name=u["name"],
                    role=role,
                    org_id=org_id,
                    team_id=team_id,
                    division_id=div_id,
                )
                print(f"  [사용자] '{u['email']}' 생성 (PW: {result['temp_password']})")
            except Exception as e:
                print(f"  [사용자] '{u['email']}' 생성 실패: {e}")

    # ── 5. 역량 ──
    for cap in data["capabilities"]:
        cap_org_id = org_map.get(cap["org"], org_id)
        try:
            await client.table("capabilities").insert({
                "type": cap["type"],
                "title": cap["title"],
                "detail": cap["detail"],
                "keywords": cap["keywords"],
                "org_id": cap_org_id,
            }).execute()
            print(f"  [역량] '{cap['title']}' 등록")
        except Exception as e:
            print(f"  [역량] '{cap['title']}' 실패: {e}")


async def main():
    if len(sys.argv) < 2:
        print("Usage: uv run python scripts/seed_org_structure.py <xlsx_path>")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    if not os.path.exists(xlsx_path):
        print(f"파일을 찾을 수 없습니다: {xlsx_path}")
        sys.exit(1)

    print(f"XLSX 파싱: {xlsx_path}")
    data = parse_xlsx(xlsx_path)

    print("\n파싱 결과:")
    print(f"  조직: {len(data['orgs'])}개")
    print(f"  본부: {len(data['divisions'])}개")
    print(f"  팀:   {len(data['teams'])}개")
    print(f"  사용자: {len(data['users'])}명")
    print(f"  역량: {len(data['capabilities'])}개")

    print("\nDB 삽입 시작...")
    await seed(data)
    print("\n완료!")


if __name__ == "__main__":
    asyncio.run(main())
